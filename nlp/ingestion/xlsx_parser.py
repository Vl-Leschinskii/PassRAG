"""Парсинг XLSX с разметкой вопросов."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from nlp.topics import TOPICS

LABEL_SHEETS = ("Для заполнения", "Вопрос и ответ ИМ и СЭД")


def _row_labels(row: tuple, header: tuple) -> list[str]:
    labels = []
    for i, topic in enumerate(TOPICS):
        col_idx = None
        for j, name in enumerate(header):
            if name and topic in str(name):
                col_idx = j
                break
        if col_idx is None:
            continue
        if col_idx < len(row) and row[col_idx] in (1, "1", True):
            labels.append(topic)
    return labels


def parse_xlsx_labels(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    questions: list[dict[str, Any]] = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name not in LABEL_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            question = str(row[0]).strip()
            if question in seen:
                continue
            seen.add(question)
            labels = _row_labels(row, header)
            item: dict[str, Any] = {"question": question, "labels": labels}
            if "Рекомендация" in [str(h) for h in header if h]:
                rec_idx = next(
                    (i for i, h in enumerate(header) if h and "Рекомендация" in str(h)),
                    None,
                )
                if rec_idx is not None and rec_idx < len(row) and row[rec_idx]:
                    item["recommendation"] = str(row[rec_idx])
            questions.append(item)

    topic_examples: dict[str, str] = {}
    if "темы вопросов" in wb.sheetnames:
        ws = wb["темы вопросов"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                topic_examples[str(row[0])] = str(row[1] or "")

    return {"questions": questions, "topic_examples": topic_examples, "topics": TOPICS}
